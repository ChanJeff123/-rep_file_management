#!/usr/bin/env python
# -*- coding: utf-8 -*-

import hashlib
import random
import openpyxl
from openpyxl import Workbook
import requests

# 使用: 将要翻译的内容放到c:\_Work\source.xlsx的Sheet1的第1列

# set baidu develop parameter
apiurl = 'http://api.fanyi.baidu.com/api/trans/vip/translate'
appid = '20190813000326115'
secretKey = 'OaD28VBAOxn7_fVntnQF'
# appid = '20191216000366578'
# secretKey = 'i3pjXzZIbB4rXcZdbKcW'
        #appid = '2018052000016305'
        #secreKey = 'rnE7cs0NwY9gQfNsTSZ'
                #appid = '20180520000163053'
        #secreKey = 'rnE7cs0NwY9gQfNsTSZF'
#         appid = '20190701000313590'
# secretKey = 'oGgUqsrGQl3GV2Oxrd_8'
#   appid = '20180115000115453'
#     secret = 'AHYyafcYLzm3q0fkfQ7z'

# 翻译内容 源语言 翻译后的语言
def translateBaidu(content, fromLang='en', toLang='zh'):
    salt = str(random.randint(32768, 65536))
    sign = appid + content + salt + secretKey
    sign = hashlib.md5(sign.encode("utf-8")).hexdigest()

    try:
        paramas = {
            'appid': appid,
            'q': content,
            'from': fromLang,
            'to': toLang,
            'salt': salt,
            'sign': sign
        }
        response = requests.get(apiurl, paramas)
        jsonResponse = response.json()  # 获得返回的结果，结果为json格式
        dst = str(jsonResponse["trans_result"]
                  [0]["dst"])  # 取得翻译后的文本结果
        return dst
    except Exception as e:
        print(e)


def excelTrans(srcFilename=r'D:\projects\爬虫项目\脚本\商标\us0.xlsx', 
               desFilename=r'D:\projects\爬虫项目\脚本\商标\us0.xlsx',
               srcSheet='Sheet',  
            #    srcColumn=1, srcRowBegin=1, srcRowEnd=11432, 
               srcColumn=1, srcRowBegin=1, srcRowEnd=4000, 
               desColumn=1, desSheet='result1'):
    wb = openpyxl.load_workbook(srcFilename)
    ws = wb[srcSheet]
    wb2 = Workbook()
    ws2 = wb2.create_sheet(title=desSheet)
    count=1
    for i in range(srcRowBegin, srcRowEnd, 1):
        print(str(count)+"\n")
        result = ws.cell(row=i, column=srcColumn).value
        if not (result is None):
            ws2.cell(row=i-srcRowBegin+1,
                     column=desColumn).value = translateBaidu(result)
        count+=1
    wb2.save(desFilename)


if __name__ == '__main__':
    print('translate begin...')
    try:
        excelTrans()
    except:
        pass
    print('ending...')
